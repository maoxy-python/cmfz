from openpyxl import Workbook
import datetime
import time

if __name__ == '__main__':
    # 创建文件对象
    wb = Workbook()

    # 获取第一个sheet
    ws = wb.active

    # 写入数字
    ws['A1'] = 42

    # 写入中文（unicode中文也可）
    ws['B1'] = "你好" + "automation test"

    # 写入多个单元格
    ws.append([1, 2, 3])

    ws['A2'] = datetime.datetime.now()  # 写入一个当前时间
    # 写入一个自定义的时间格式
    ws['A3'] = time.strftime('%Y{y}%m{m}%d{d}').format(y='年', m='月', d='日')

    wb.save("e:\\sample.xlsx")